"""XLSX adapter — openpyxl, sheet header + tab-delimited 단일 body.

각 sheet 를 `## Sheet: <name>` 헤더로 구분, row 는 tab-separated.
간단 default. sheet 별 분리 인덱싱은 SPEC 정제 후 확장.
"""
from __future__ import annotations

from pathlib import Path

from core.knowledge.adapters._base import ExtractedDocument


class XlsxAdapter:
    extensions: tuple[str, ...] = (".xlsx",)

    def extract(self, path: Path) -> ExtractedDocument:
        try:
            from openpyxl import load_workbook
        except ImportError as e:
            raise RuntimeError(
                "openpyxl required for xlsx adapter — `uv add openpyxl`"
            ) from e

        wb = load_workbook(str(path), read_only=True, data_only=True)
        parts: list[str] = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            parts.append(f"## Sheet: {sheet_name}")
            for row in ws.iter_rows(values_only=True):
                cells = ["" if v is None else str(v) for v in row]
                if any(cells):
                    parts.append("\t".join(cells))
            parts.append("")
        body = "\n".join(parts).strip()
        return ExtractedDocument(
            body_text=body,
            extraction_meta={
                "sheet_count": len(wb.sheetnames),
                "char_count": len(body),
            },
        )
